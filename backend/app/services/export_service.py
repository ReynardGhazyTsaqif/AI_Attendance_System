import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from sqlalchemy.orm import Session
from datetime import datetime
from calendar import monthrange
import io

from app.models.attendance_model import Attendance
from app.models.user_model import User

def export_attendance_excel(db: Session, year: int, month: int) -> bytes:
    """Export data absensi bulanan ke Excel."""
    _, days = monthrange(year, month)
    start = datetime(year, month, 1)
    end = datetime(year, month, days, 23, 59, 59)

    records = db.query(Attendance, User).join(
        User, Attendance.user_id == User.id
    ).filter(
        Attendance.check_in_time.between(start, end)
    ).order_by(User.name, Attendance.check_in_time).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Absensi {year}-{month:02d}"

    # Style
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"LAPORAN ABSENSI — {datetime(year, month, 1).strftime('%B %Y').upper()}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center
    ws.row_dimensions[1].height = 30

    # Header row
    headers = ["No", "Nama", "Tanggal", "Check In", "Check Out", "Status", "Terlambat (menit)", "Confidence (%)", "Lokasi"]
    ws.row_dimensions[3].height = 22
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    # Data rows
    status_colors = {
        "hadir": "D1FAE5",
        "terlambat": "FEF3C7",
        "tidak_hadir": "FEE2E2",
        "izin": "DBEAFE",
        "sakit": "EDE9FE",
        "pulang_cepat": "FFE4E6"
    }

    for row_idx, (attendance, user) in enumerate(records, start=4):
        row_data = [
            row_idx - 3,
            user.name,
            attendance.check_in_time.strftime("%d/%m/%Y") if attendance.check_in_time else "-",
            attendance.check_in_time.strftime("%H:%M:%S") if attendance.check_in_time else "-",
            attendance.check_out_time.strftime("%H:%M:%S") if attendance.check_out_time else "-",
            attendance.status.replace("_", " ").title(),
            attendance.late_minutes or 0,
            round(attendance.confidence_score, 1) if attendance.confidence_score else "-",
            f"{attendance.location_lat}, {attendance.location_lng}" if attendance.location_lat else "-"
        ]

        fill_color = status_colors.get(attendance.status, "FFFFFF")
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin
            cell.alignment = center
            if col == 6:  # Status column
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Auto column width
    col_widths = [5, 25, 12, 12, 12, 15, 18, 15, 25]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_attendance_pdf(db: Session, year: int, month: int) -> bytes:
    """Export data absensi bulanan ke PDF."""
    _, days = monthrange(year, month)
    start = datetime(year, month, 1)
    end = datetime(year, month, days, 23, 59, 59)

    records = db.query(Attendance, User).join(
        User, Attendance.user_id == User.id
    ).filter(
        Attendance.check_in_time.between(start, end)
    ).order_by(User.name, Attendance.check_in_time).all()

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle('title', parent=styles['Title'], fontSize=16, spaceAfter=12)
    elements.append(Paragraph(f"Laporan Absensi — {datetime(year, month, 1).strftime('%B %Y')}", title_style))
    elements.append(Spacer(1, 0.3*cm))

    # Table data
    table_data = [["No", "Nama", "Tanggal", "Check In", "Check Out", "Status", "Terlambat", "Confidence"]]

    for idx, (attendance, user) in enumerate(records, 1):
        table_data.append([
            str(idx),
            user.name,
            attendance.check_in_time.strftime("%d/%m/%Y") if attendance.check_in_time else "-",
            attendance.check_in_time.strftime("%H:%M") if attendance.check_in_time else "-",
            attendance.check_out_time.strftime("%H:%M") if attendance.check_out_time else "-",
            attendance.status.replace("_", " ").title(),
            f"{attendance.late_minutes or 0} mnt",
            f"{round(attendance.confidence_score, 1) if attendance.confidence_score else '-'}%"
        ])

    # Table style
    table = Table(table_data, colWidths=[1*cm, 5*cm, 3*cm, 2.5*cm, 2.5*cm, 3*cm, 2.5*cm, 2.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWHEIGHT', (0, 0), (-1, -1), 20),
    ]))

    elements.append(table)
    doc.build(elements)
    return output.getvalue()